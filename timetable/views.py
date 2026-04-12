from rest_framework import viewsets, status, generics, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.db import transaction
from django.db.models import Q, Count, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.core.cache import cache
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
# FIX: timedelta was used in WebSocketTokenView but never imported
from datetime import datetime, date, timedelta
import json

from .models import (
    AcademicYear, Semester, Department, Programme, Stage, Unit,
    Intake, IntakeUnit, Lecturer, Room, TimeSlot, TimetableEntry,
    ConflictLog, ScheduleAudit,
)
from .serializers import (
    AcademicYearSerializer, SemesterSerializer, DepartmentSerializer,
    ProgrammeSerializer, StageSerializer, UnitSerializer,
    IntakeSerializer, LecturerSerializer, RoomSerializer,
    TimeSlotSerializer, TimetableEntrySerializer, ConflictLogSerializer,
)
from .permissions import IsAdminUser, IsCoordinator, IsLecturer


# ============== Helper: safe cache invalidation ==============

def _clear_timetable_cache(semester_id=None):
    """
    Safely delete timetable cache keys.
    cache.delete_pattern() only exists in django-redis.  We use explicit keys
    so the code works with any cache backend.
    """
    keys = [
        f"master_timetable_{semester_id}" if semester_id else "master_timetable_all",
    ]
    for key in keys:
        cache.delete(key)


# ============== ViewSets for CRUD Operations ==============

class AcademicYearViewSet(viewsets.ModelViewSet):
    """Manage academic years."""
    queryset = AcademicYear.objects.filter(is_deleted=False)
    serializer_class = AcademicYearSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['is_current']
    search_fields = ['year', 'name']

    @action(detail=False, methods=['get'])
    def current(self, request):
        """Get current academic year."""
        current = AcademicYear.objects.filter(is_current=True, is_deleted=False).first()
        if current:
            serializer = self.get_serializer(current)
            return Response(serializer.data)
        return Response({'detail': 'No current academic year set'}, status=404)


class SemesterViewSet(viewsets.ModelViewSet):
    """Manage semesters."""
    queryset = Semester.objects.filter(is_deleted=False)
    serializer_class = SemesterSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['academic_year', 'semester_type', 'is_active']
    search_fields = ['name']

    @action(detail=False, methods=['get'])
    def active(self, request):
        """Get active semester."""
        semester = Semester.objects.filter(is_active=True, is_deleted=False).first()
        if semester:
            serializer = self.get_serializer(semester)
            return Response(serializer.data)
        return Response({'detail': 'No active semester found'}, status=404)

    @action(detail=False, methods=['post'])
    def create_three_semesters(self, request):
        """Create all three semesters for an academic year."""
        academic_year_id = request.data.get('academic_year_id')
        try:
            academic_year = AcademicYear.objects.get(id=academic_year_id)
        except AcademicYear.DoesNotExist:
            return Response({'error': 'Academic year not found'}, status=404)

        semesters_data = {
            'JAN_APR': {'start_month': 1, 'end_month': 4},
            'MAY_AUG': {'start_month': 5, 'end_month': 8},
            'SEP_DEC': {'start_month': 9, 'end_month': 12},
        }

        created_semesters = []
        for sem_type, dates in semesters_data.items():
            semester, created = Semester.objects.get_or_create(
                academic_year=academic_year,
                semester_type=sem_type,
                defaults={
                    'name': f"{academic_year.year} - {dict(Semester.SEMESTER_CHOICES)[sem_type]}",
                    'start_date': date(academic_year.year, dates['start_month'], 1),
                    'end_date': date(academic_year.year, dates['end_month'], 28),
                    'registration_deadline': date(academic_year.year, dates['start_month'], 15),
                    # FIX: end_month=4 has 30 days; 30 is fine for APR/JUN/SEP/NOV but
                    # using 28 is a safe floor — administrators should adjust per year.
                    'add_drop_deadline': date(academic_year.year, dates['start_month'], 21),
                    'withdrawal_deadline': date(academic_year.year, dates['start_month'] + 1, 15),
                    'teaching_weeks': 14,
                },
            )
            created_semesters.append(semester)

        serializer = self.get_serializer(created_semesters, many=True)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class DepartmentViewSet(viewsets.ModelViewSet):
    """Manage departments."""
    queryset = Department.objects.filter(is_deleted=False)
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['name', 'code']


class ProgrammeViewSet(viewsets.ModelViewSet):
    """Manage programmes."""
    queryset = Programme.objects.filter(is_deleted=False)
    serializer_class = ProgrammeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['department', 'programme_type']
    search_fields = ['name', 'code']


class StageViewSet(viewsets.ModelViewSet):
    """Manage stages."""
    queryset = Stage.objects.filter(is_deleted=False)
    serializer_class = StageSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['programme']


class UnitViewSet(viewsets.ModelViewSet):
    """Manage units."""
    queryset = Unit.objects.filter(is_deleted=False)
    serializer_class = UnitSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['stage', 'unit_type']
    search_fields = ['code', 'name']

    @action(detail=True, methods=['get'])
    def qualified_lecturers(self, request, pk=None):
        """Get lecturers qualified to teach this unit."""
        unit = self.get_object()
        lecturers = unit.qualified_lecturers.filter(is_active=True)
        serializer = LecturerSerializer(lecturers, many=True)
        return Response(serializer.data)


class IntakeViewSet(viewsets.ModelViewSet):
    """Manage intakes."""
    queryset = Intake.objects.filter(is_deleted=False)
    serializer_class = IntakeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['programme', 'stage', 'academic_year', 'is_active']
    search_fields = ['name']

    @action(detail=True, methods=['post'])
    def assign_units(self, request, pk=None):
        """Assign units to intake for a semester."""
        intake = self.get_object()
        units_data = request.data.get('units', [])
        semester_id = request.data.get('semester_id')

        if not semester_id:
            return Response({'error': 'semester_id required'}, status=400)

        semester = get_object_or_404(Semester, id=semester_id)

        with transaction.atomic():
            IntakeUnit.objects.filter(intake=intake, semester=semester).delete()
            for unit_data in units_data:
                IntakeUnit.objects.create(
                    intake=intake,
                    unit_id=unit_data['unit_id'],
                    semester=semester,
                    is_mandatory=unit_data.get('is_mandatory', True),
                )

        return Response({'status': 'Units assigned successfully'})


class LecturerViewSet(viewsets.ModelViewSet):
    """Manage lecturers."""
    queryset = Lecturer.objects.filter(is_deleted=False)
    serializer_class = LecturerSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['department', 'lecturer_type', 'is_active']
    search_fields = ['first_name', 'last_name', 'staff_id', 'email']


class RoomViewSet(viewsets.ModelViewSet):
    """Manage rooms."""
    queryset = Room.objects.filter(is_deleted=False)
    serializer_class = RoomSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['room_type', 'building']
    search_fields = ['code', 'name']


class TimeSlotViewSet(viewsets.ReadOnlyModelViewSet):
    """View available time slots."""
    queryset = TimeSlot.objects.all()
    serializer_class = TimeSlotSerializer
    permission_classes = [IsAuthenticated]


class TimetableEntryViewSet(viewsets.ModelViewSet):
    """Manage timetable entries."""
    queryset = TimetableEntry.objects.filter(is_deleted=False)
    serializer_class = TimetableEntrySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['semester', 'intake', 'lecturer', 'room', 'status', 'day', 'week_number']
    search_fields = ['unit__code', 'lecturer__first_name']


class ConflictViewSet(viewsets.ModelViewSet):
    """Manage conflicts."""
    queryset = ConflictLog.objects.filter(is_deleted=False)
    serializer_class = ConflictLogSerializer
    permission_classes = [IsAuthenticated, IsCoordinator]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['semester', 'conflict_type', 'resolution_status']

    @action(detail=True, methods=['post'])
    def resolve(self, request, pk=None):
        """Resolve a conflict."""
        conflict = self.get_object()
        resolution = request.data.get('resolution')

        conflict.resolution_status = request.data.get('method', 'MANUAL')
        conflict.resolved_by = request.user
        conflict.resolved_at = timezone.now()
        conflict.proposed_solution = {'resolution': resolution}
        conflict.save()

        return Response({'status': 'Conflict resolved'})


# ============== Timetable Views ==============

class MasterTimetableView(APIView):
    """Get master timetable grid."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        semester_id = request.query_params.get('semester_id')

        if semester_id:
            semester = get_object_or_404(Semester, id=semester_id)
        else:
            semester = Semester.objects.filter(is_active=True).first()

        if not semester:
            return Response({'error': 'No active semester'}, status=404)

        cache_key = f"master_timetable_{semester.id}"
        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data)

        entries = TimetableEntry.objects.filter(
            semester=semester,
            status='PUBLISHED',
            is_deleted=False,
        ).select_related('unit', 'lecturer', 'room', 'intake', 'time_slot')

        days = ['MON', 'TUE', 'WED', 'THU', 'FRI']
        time_slots = TimeSlot.objects.all().order_by('order')

        grid = {day: {slot.slot_id: [] for slot in time_slots} for day in days}

        for entry in entries:
            grid[entry.day][entry.time_slot.slot_id].append({
                'id': str(entry.id),
                'unit_code': entry.unit.code,
                'unit_name': entry.unit.name,
                'lecturer': entry.lecturer.full_name,
                'lecturer_id': str(entry.lecturer.id),
                'room': entry.room.name,
                'room_code': entry.room.code,
                'intake': entry.intake.name,
                'intake_id': str(entry.intake.id),
            })

        slot_labels = {slot.slot_id: slot.get_slot_id_display() for slot in time_slots}

        response_data = {
            'semester': semester.name,
            'semester_id': str(semester.id),
            'days': days,
            'slot_labels': slot_labels,
            'grid': grid,
            'total_weeks': semester.teaching_weeks,
        }

        cache.set(cache_key, response_data, timeout=300)
        return Response(response_data)


class PersonalTimetableView(APIView):
    """Get personal timetable for logged-in lecturer."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'lecturer_profile'):
            return Response({'error': 'User is not a lecturer'}, status=403)

        lecturer = request.user.lecturer_profile
        semester_id = request.query_params.get('semester_id')

        if semester_id:
            semester = get_object_or_404(Semester, id=semester_id)
        else:
            semester = Semester.objects.filter(is_active=True).first()

        if not semester:
            return Response({'error': 'No active semester'}, status=404)

        cache_key = f"personal_timetable_{lecturer.id}_{semester.id}"
        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data)

        entries = TimetableEntry.objects.filter(
            semester=semester,
            lecturer=lecturer,
            status='PUBLISHED',
            is_deleted=False,
        ).select_related('unit', 'room', 'intake', 'time_slot').order_by('day', 'time_slot__order')

        days = ['MON', 'TUE', 'WED', 'THU', 'FRI']
        time_slots = TimeSlot.objects.all().order_by('order')

        grid = {day: {slot.slot_id: None for slot in time_slots} for day in days}
        total_slots = 0

        for entry in entries:
            grid[entry.day][entry.time_slot.slot_id] = {
                'unit_code': entry.unit.code,
                'unit_name': entry.unit.name,
                'room': entry.room.name,
                'room_code': entry.room.code,
                'intake': entry.intake.name,
                'intake_id': str(entry.intake.id),
            }
            total_slots += 1

        slot_labels = {slot.slot_id: slot.get_slot_id_display() for slot in time_slots}

        total_entries = TimetableEntry.objects.filter(
            semester=semester,
            lecturer=lecturer,
            status='PUBLISHED',
            is_deleted=False,
        ).count()

        response_data = {
            'lecturer_id': str(lecturer.id),
            'lecturer_name': lecturer.full_name,
            'lecturer_type': lecturer.get_lecturer_type_display(),
            'available_days': lecturer.get_available_days(),
            'semester': semester.name,
            'semester_id': str(semester.id),
            'days': days,
            'slot_labels': slot_labels,
            'grid': grid,
            'total_hours': total_slots * 2,
            'max_hours_per_week': lecturer.max_hours_per_week,
            'hours_remaining': lecturer.max_hours_per_week - (total_slots * 2),
            'total_classes_semester': total_entries,
        }

        cache.set(cache_key, response_data, timeout=300)
        return Response(response_data)


# ============== Scheduling Views ==============

class GenerateTimetableView(APIView):
    """Generate timetable using scheduler."""
    permission_classes = [IsAuthenticated, IsCoordinator]

    def post(self, request):
        semester_id = request.data.get('semester_id')
        algorithm = request.data.get('algorithm', 'HYBRID')

        if not semester_id:
            return Response({'error': 'semester_id required'}, status=400)

        semester = get_object_or_404(Semester, id=semester_id)

        # Clear existing draft entries
        TimetableEntry.objects.filter(semester=semester, status='DRAFT').delete()

        self._create_sample_timetable(semester)

        # FIX: use explicit key deletion instead of cache.delete_pattern
        _clear_timetable_cache(semester_id=semester.id)

        return Response(
            {
                'status': 'success',
                'message': f'Timetable generated for {semester.name}',
                'algorithm': algorithm,
            },
            status=status.HTTP_202_ACCEPTED,
        )

    def _create_sample_timetable(self, semester):
        """Create sample timetable entries for testing (skips model-level clean)."""
        import random

        intakes = list(Intake.objects.filter(is_active=True)[:3])
        rooms = list(Room.objects.filter(is_active=True)[:3])
        time_slots = list(TimeSlot.objects.all())
        days = ['MON', 'TUE', 'WED', 'THU', 'FRI']

        for week in range(1, 5):
            for intake in intakes:
                intake_units = IntakeUnit.objects.filter(intake=intake, semester=semester)[:3]
                for iu in intake_units:
                    if not iu.unit.qualified_lecturers.exists():
                        continue

                    lecturer = iu.unit.qualified_lecturers.first()
                    room = random.choice(rooms) if rooms else None
                    day = random.choice(days)
                    time_slot = random.choice(time_slots) if time_slots else None

                    if room and time_slot:
                        # Use update_or_create to avoid unique-constraint errors on re-runs
                        TimetableEntry.objects.update_or_create(
                            semester=semester,
                            intake=intake,
                            unit=iu.unit,
                            lecturer=lecturer,
                            day=day,
                            time_slot=time_slot,
                            week_number=week,
                            defaults={'room': room, 'status': 'DRAFT'},
                        )


class PublishTimetableView(APIView):
    """Publish timetable and notify lecturers."""
    permission_classes = [IsAuthenticated, IsCoordinator]

    def post(self, request):
        semester_id = request.data.get('semester_id')

        if not semester_id:
            return Response({'error': 'semester_id required'}, status=400)

        semester = get_object_or_404(Semester, id=semester_id)

        with transaction.atomic():
            updated = TimetableEntry.objects.filter(
                semester=semester,
                status='DRAFT',
                is_deleted=False,
            ).update(status='PUBLISHED', published_at=timezone.now())

        # FIX: explicit cache key deletion
        _clear_timetable_cache(semester_id=semester.id)

        return Response({
            'published': updated,
            'message': f'Published {updated} classes for {semester.name}',
        })


# ============== Export Views ==============

class ExportMasterPDFView(APIView):
    """Export master timetable as an HTML file (printable to PDF)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        semester_id = request.query_params.get('semester_id')

        if semester_id:
            semester = get_object_or_404(Semester, id=semester_id)
        else:
            semester = Semester.objects.filter(is_active=True).first()

        if not semester:
            return Response({'error': 'No semester found'}, status=404)

        entries = TimetableEntry.objects.filter(
            semester=semester,
            status='PUBLISHED',
            is_deleted=False,
        ).select_related('unit', 'lecturer', 'room', 'intake', 'time_slot')

        days = ['MON', 'TUE', 'WED', 'THU', 'FRI']
        time_slots = TimeSlot.objects.all().order_by('order')

        html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Master Timetable - {semester.name}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        h1 {{ color: #333; text-align: center; }}
        h2 {{ color: #666; text-align: center; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
        th {{ background-color: #4CAF50; color: white; padding: 12px; text-align: center; }}
        td {{ border: 1px solid #ddd; padding: 8px; vertical-align: top; }}
        .class-cell {{ background-color: #f9f9f9; }}
        .empty-cell {{ background-color: #fff; text-align: center; color: #999; }}
    </style>
</head>
<body>
    <h1>MASTER TIMETABLE</h1>
    <h2>{semester.name}</h2>
    <h3>Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}</h3>
    <table>
        <thead>
            <tr>
                <th>Day / Time</th>"""

        for slot in time_slots:
            html_content += f"<th>{slot.get_slot_id_display()}</th>"

        html_content += "</tr></thead><tbody>"

        for day in days:
            html_content += f"<tr><td><strong>{day.title()}</strong></td>"
            for slot in time_slots:
                entry = entries.filter(day=day, time_slot=slot).first()
                if entry:
                    html_content += (
                        f'<td class="class-cell">'
                        f'<strong>{entry.unit.code}</strong><br>'
                        f'{entry.lecturer.short_name}<br>'
                        f'{entry.room.code}<br>'
                        f'<small>{entry.intake.name}</small>'
                        f'</td>'
                    )
                else:
                    html_content += '<td class="empty-cell">—</td>'
            html_content += "</tr>"

        html_content += "</tbody></table></body></html>"

        response = HttpResponse(html_content, content_type='text/html')
        response['Content-Disposition'] = (
            f'inline; filename="master_timetable_{semester.name}.html"'
        )
        return response


class ExportPersonalPDFView(APIView):
    """Export personal timetable as an HTML file (printable to PDF)."""
    permission_classes = [IsAuthenticated]

    def get(self, request, lecturer_id):
        lecturer = get_object_or_404(Lecturer, id=lecturer_id)
        semester = Semester.objects.filter(is_active=True).first()

        if not semester:
            return Response({'error': 'No semester found'}, status=404)

        entries = TimetableEntry.objects.filter(
            semester=semester,
            lecturer=lecturer,
            status='PUBLISHED',
            is_deleted=False,
        ).select_related('unit', 'room', 'intake', 'time_slot')

        days = ['MON', 'TUE', 'WED', 'THU', 'FRI']
        time_slots = TimeSlot.objects.all().order_by('order')

        html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Personal Timetable - {lecturer.full_name}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        h1, h2 {{ text-align: center; }}
        .info {{ text-align: center; margin: 20px 0; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
        th {{ background-color: #2196F3; color: white; padding: 12px; }}
        td {{ border: 1px solid #ddd; padding: 8px; }}
    </style>
</head>
<body>
    <h1>PERSONAL TIMETABLE</h1>
    <h2>{lecturer.full_name}</h2>
    <div class="info">
        <p>Department: {lecturer.department.name}</p>
        <p>Staff ID: {lecturer.staff_id}</p>
        <p>Semester: {semester.name}</p>
    </div>
    <table>
        <thead><tr><th>Day / Time</th>"""

        for slot in time_slots:
            html_content += f"<th>{slot.get_slot_id_display()}</th>"

        html_content += "</tr></thead><tbody>"

        for day in days:
            html_content += f"<tr><td><strong>{day.title()}</strong></td>"
            for slot in time_slots:
                entry = entries.filter(day=day, time_slot=slot).first()
                if entry:
                    html_content += (
                        f'<td>'
                        f'<strong>{entry.unit.code}</strong><br>'
                        f'{entry.unit.name}<br>'
                        f'{entry.room.code}<br>'
                        f'<small>{entry.intake.name}</small>'
                        f'</td>'
                    )
                else:
                    html_content += '<td style="text-align:center; color:#999;">Free</td>'
            html_content += "</tr>"

        html_content += "</tbody></table></body></html>"

        response = HttpResponse(html_content, content_type='text/html')
        response['Content-Disposition'] = (
            f'inline; filename="timetable_{lecturer.staff_id}.html"'
        )
        return response


class ExportExcelView(APIView):
    """Export timetable to Excel."""
    permission_classes = [IsAuthenticated, IsCoordinator]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        semester_id = request.query_params.get('semester_id')

        if semester_id:
            semester = get_object_or_404(Semester, id=semester_id)
        else:
            semester = Semester.objects.filter(is_active=True).first()

        if not semester:
            return Response({'error': 'No semester found'}, status=404)

        wb = openpyxl.Workbook()

        intakes = Intake.objects.filter(
            intakeunit__semester=semester,
            is_active=True,
            is_deleted=False,
        ).distinct()

        for intake in intakes:
            ws = wb.create_sheet(title=intake.name[:31])
            time_slots = TimeSlot.objects.all().order_by('order')
            headers = ['Day / Time'] + [slot.get_slot_id_display() for slot in time_slots]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            entries = TimetableEntry.objects.filter(
                semester=semester,
                intake=intake,
                status='PUBLISHED',
                is_deleted=False,
            ).select_related('unit', 'lecturer', 'room', 'time_slot')

            days = ['MON', 'TUE', 'WED', 'THU', 'FRI']
            row = 2

            for day in days:
                ws.cell(row=row, column=1, value=day.title()).font = Font(bold=True)
                for col, slot in enumerate(time_slots, 2):
                    entry = entries.filter(day=day, time_slot=slot).first()
                    if entry:
                        value = f"{entry.unit.code}\n{entry.lecturer.short_name}\n{entry.room.code}"
                        c = ws.cell(row=row, column=col, value=value)
                        c.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    else:
                        ws.cell(row=row, column=col, value="—").alignment = Alignment(horizontal='center')
                row += 1

            for column in ws.columns:
                max_length = max(
                    (len(str(cell.value)) for cell in column if cell.value), default=0
                )
                ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

        # Remove the default empty sheet created by openpyxl
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="timetable_{semester.name}.xlsx"'
        )
        wb.save(response)
        return response


# ============== Dashboard Views ==============

class DashboardStatsView(APIView):
    """Get dashboard statistics."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        semester = Semester.objects.filter(is_active=True).first()

        stats = {
            'total_lecturers': Lecturer.objects.filter(is_active=True, is_deleted=False).count(),
            'total_rooms': Room.objects.filter(is_active=True, is_deleted=False).count(),
            'total_intakes': Intake.objects.filter(is_active=True, is_deleted=False).count(),
            'total_units': Unit.objects.filter(is_active=True, is_deleted=False).count(),
            'total_departments': Department.objects.filter(is_active=True, is_deleted=False).count(),
            'total_programmes': Programme.objects.filter(is_active=True, is_deleted=False).count(),
        }

        if semester:
            published_classes = TimetableEntry.objects.filter(
                semester=semester, status='PUBLISHED', is_deleted=False
            ).count()
            total_slots_possible = 5 * 3 * semester.teaching_weeks
            utilization = (published_classes / total_slots_possible * 100) if total_slots_possible > 0 else 0

            stats['current_semester'] = {
                'id': str(semester.id),
                'name': semester.name,
                'published_classes': published_classes,
                'draft_classes': TimetableEntry.objects.filter(
                    semester=semester, status='DRAFT', is_deleted=False
                ).count(),
                'cancelled_classes': TimetableEntry.objects.filter(
                    semester=semester, status='CANCELLED', is_deleted=False
                ).count(),
                'utilization_rate': round(utilization, 2),
                'current_week': semester.current_week,
                'weeks_remaining': semester.weeks_remaining,
            }

        lecturers = Lecturer.objects.filter(is_active=True, is_deleted=False)[:10]
        stats['lecturer_workload'] = [
            {
                'name': lec.short_name,
                'current_hours': lec.get_current_workload(semester) if semester else 0,
                'max_hours': lec.max_hours_per_week,
            }
            for lec in lecturers
        ]

        recent_audits = ScheduleAudit.objects.select_related(
            'timetable_entry', 'changed_by'
        ).order_by('-created_at')[:10]

        stats['recent_activities'] = [
            {
                'action': audit.get_action_display(),
                'entry': str(audit.timetable_entry),
                'user': audit.changed_by.username if audit.changed_by else 'System',
                'time': audit.created_at.strftime('%Y-%m-%d %H:%M'),
            }
            for audit in recent_audits
        ]

        return Response(stats)


class LecturerDashboardView(APIView):
    """Get dashboard for logged-in lecturer."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'lecturer_profile'):
            return Response({'error': 'User is not a lecturer'}, status=403)

        lecturer = request.user.lecturer_profile
        semester = Semester.objects.filter(is_active=True).first()

        if not semester:
            return Response({'error': 'No active semester'}, status=404)

        current_week = semester.current_week
        current_week_entries = TimetableEntry.objects.filter(
            semester=semester,
            lecturer=lecturer,
            week_number=current_week,
            status='PUBLISHED',
            is_deleted=False,
        ).count()

        total_entries = TimetableEntry.objects.filter(
            semester=semester,
            lecturer=lecturer,
            status='PUBLISHED',
            is_deleted=False,
        ).count()

        today = timezone.now().date()
        day_map = {'MON': 0, 'TUE': 1, 'WED': 2, 'THU': 3, 'FRI': 4}

        entries = TimetableEntry.objects.filter(
            semester=semester,
            lecturer=lecturer,
            week_number=current_week,
            status='PUBLISHED',
            is_deleted=False,
        ).select_related('unit', 'room', 'intake', 'time_slot')

        upcoming = []
        for entry in entries:
            days_ahead = day_map.get(entry.day, 0) - today.weekday()
            if days_ahead < 0:
                days_ahead += 7
            class_date = today + timedelta(days=days_ahead)
            upcoming.append({
                'unit': entry.unit.code,
                'unit_name': entry.unit.name,
                'day': entry.day,
                'time': entry.time_slot.get_slot_id_display(),
                'room': entry.room.code,
                'intake': entry.intake.name,
                'date': class_date.strftime('%Y-%m-%d'),
            })

        return Response({
            'lecturer': {
                'name': lecturer.full_name,
                'staff_id': lecturer.staff_id,
                'department': lecturer.department.name,
                'type': lecturer.get_lecturer_type_display(),
            },
            'semester': semester.name,
            'current_week': current_week,
            'current_week_classes': current_week_entries,
            'total_weekly_hours': total_entries * 2,
            'max_weekly_hours': lecturer.max_hours_per_week,
            'hours_remaining': lecturer.max_hours_per_week - (total_entries * 2),
            'upcoming_classes': upcoming[:10],
            'statistics': {
                'total_classes_semester': total_entries,
                'total_units': lecturer.qualified_units.count(),
                'weeks_remaining': semester.weeks_remaining,
            },
        })


class WebSocketTokenView(APIView):
    """Generate token for WebSocket authentication."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import jwt
        from django.conf import settings

        # FIX: timedelta is now imported at the top of the file
        token = jwt.encode(
            {
                'user_id': str(request.user.id),
                'username': request.user.username,
                'exp': timezone.now() + timedelta(hours=1),
            },
            settings.SECRET_KEY,
            algorithm='HS256',
        )

        return Response({'token': token})