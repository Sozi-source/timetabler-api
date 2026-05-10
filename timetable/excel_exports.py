"""
timetable/excel_exports.py
==========================
Excel (.xlsx) export helpers for master, cohort, and trainer timetables.

Public API consumed by views.py
--------------------------------
_make_workbook(sheets: list[dict]) -> bytes
    Build an xlsx workbook from one or more sheet definitions.
    Each sheet dict has keys:
        sheet_name : str
        title      : str
        subtitle   : str
        days       : list[str]
        periods    : list[dict]   — each has: id, label, start, end
        grid       : dict         — grid[day][period_id] = entry | list[entry] | None
        is_master  : bool

_xlsx_response(data: bytes, filename: str) -> HttpResponse
    Wrap raw xlsx bytes in a Django download response.

Internal helpers (also imported by views.py for the unified export endpoint)
-----------------------------------------------------------------------------
_period_list(periods)      — normalise period dicts
_su_to_dict(su)            — alias kept for back-compat (unused in new flow)
"""

from __future__ import annotations

import io
from typing import Any

from django.http import HttpResponse


# ─────────────────────────────────────────────────────────────────────────────
# Openpyxl helpers
# ─────────────────────────────────────────────────────────────────────────────

def _xl():
    """Lazy import so the module loads even when openpyxl isn't installed."""
    import openpyxl
    return openpyxl


def _styles():
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        GradientFill,
        PatternFill,
        Side,
    )
    return Alignment, Border, Font, PatternFill, Side


# ── Colour constants ──────────────────────────────────────────────────────────
C_HEADER_BG  = "16213E"    # dark navy
C_HEADER_FG  = "FFFFFF"
C_DAY_BG     = "EEF2FF"    # pale indigo
C_CELL_BG    = "F8F9FF"
C_ALT_BG     = "F0F4FF"
C_ENTRY_BG   = "DDE3FF"
C_BORDER     = "C8CEE8"
C_CODE_FG    = "16213E"
C_DET_FG     = "5A6180"
C_TITLE_FG   = "16213E"
C_SUB_FG     = "5A6180"


# ─────────────────────────────────────────────────────────────────────────────
# Period helpers
# ─────────────────────────────────────────────────────────────────────────────

def _period_list(periods: list[dict]) -> list[dict]:
    """Normalise and return the period list (identity for dicts)."""
    return list(periods)


def _su_to_dict(su) -> dict:
    """
    Back-compat shim.  In the old flow, ScheduledUnit ORM objects were passed
    directly.  In the new flow, pre-serialised dicts are used.  This handles
    both cases.
    """
    if isinstance(su, dict):
        return su
    return {
        "unit_code":    su.curriculum_unit.code,
        "unit_name":    su.curriculum_unit.name,
        "trainer_name": su.trainer.short_name if su.trainer else "",
        "room_code":    su.room.code if su.room else "",
        "cohort_name":  su.cohort.name if su.cohort else "",
    }


# ─────────────────────────────────────────────────────────────────────────────
# Cell content formatter
# ─────────────────────────────────────────────────────────────────────────────

def _cell_text(item: Any, is_master: bool) -> str:
    """Convert a grid cell value to a multi-line string for xlsx."""
    if item is None or item == [] or item == {}:
        return "–"
    if isinstance(item, list):
        chunks = []
        for e in item:
            e = _su_to_dict(e)
            line = f"{e.get('unit_code','')}  {e.get('unit_name','')}"
            detail_parts = [
                e.get("trainer_name", ""),
                e.get("room_code", ""),
                e.get("cohort_name", ""),
            ]
            detail = "  ·  ".join(p for p in detail_parts if p)
            chunks.append(f"{line}\n{detail}" if detail else line)
        return "\n\n".join(chunks)
    e = _su_to_dict(item)
    line = f"{e.get('unit_code','')}  {e.get('unit_name','')}"
    detail_parts = [e.get("trainer_name", ""), e.get("room_code", "")]
    detail = "  ·  ".join(p for p in detail_parts if p)
    return f"{line}\n{detail}" if detail else line


# ─────────────────────────────────────────────────────────────────────────────
# Core workbook builder
# ─────────────────────────────────────────────────────────────────────────────

def _make_workbook(sheets: list[dict]) -> bytes:
    """
    Build an xlsx workbook and return raw bytes.

    ``sheets`` is a list of dicts, one per worksheet:
        sheet_name, title, subtitle, days, periods, grid, is_master
    """
    openpyxl = _xl()
    Alignment, Border, Font, PatternFill, Side = _styles()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)           # remove default empty sheet

    for sheet_def in sheets:
        _build_sheet(wb, sheet_def, Alignment, Border, Font, PatternFill, Side)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_sheet(wb, sheet_def, Alignment, Border, Font, PatternFill, Side):
    sheet_name = sheet_def.get("sheet_name", "Timetable")[:31]
    title      = sheet_def.get("title", "")
    subtitle   = sheet_def.get("subtitle", "")
    days       = sheet_def.get("days", [])
    periods    = _period_list(sheet_def.get("periods", []))
    grid       = sheet_def.get("grid", {})
    is_master  = sheet_def.get("is_master", False)

    ws = wb.create_sheet(title=sheet_name)

    # ── Helper style factories ─────────────────────────────────────────────
    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _border(color=C_BORDER):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def _font(bold=False, size=9, color=C_CODE_FG, italic=False):
        return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

    def _align(wrap=True, halign="left", valign="top"):
        return Alignment(wrap_text=wrap, horizontal=halign, vertical=valign)

    # ── Title rows (rows 1-2) ──────────────────────────────────────────────
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 14

    title_cell          = ws.cell(row=1, column=1, value=title)
    title_cell.font     = _font(bold=True, size=13, color=C_TITLE_FG)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    sub_cell            = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font       = _font(size=9, color=C_SUB_FG, italic=True)
    sub_cell.alignment  = Alignment(horizontal="left", vertical="center")

    if periods:
        max_col = 1 + len(periods)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)

    # ── Header row (row 3) ─────────────────────────────────────────────────
    hdr_row = 3
    ws.row_dimensions[hdr_row].height = 28

    day_hdr = ws.cell(row=hdr_row, column=1, value="Day")
    day_hdr.font      = _font(bold=True, size=9, color=C_HEADER_FG)
    day_hdr.fill      = _fill(C_HEADER_BG)
    day_hdr.alignment = _align(wrap=False, halign="center", valign="center")
    day_hdr.border    = _border("444444")
    ws.column_dimensions["A"].width = 12

    for ci, period in enumerate(periods, start=2):
        col_letter = ws.cell(row=hdr_row, column=ci).column_letter
        ws.column_dimensions[col_letter].width = 22
        label = f"{period['label']}\n{period['start'][:5]}–{period['end'][:5]}"
        cell  = ws.cell(row=hdr_row, column=ci, value=label)
        cell.font      = _font(bold=True, size=8, color=C_HEADER_FG)
        cell.fill      = _fill(C_HEADER_BG)
        cell.alignment = _align(wrap=True, halign="center", valign="center")
        cell.border    = _border("444444")

    # ── Data rows ──────────────────────────────────────────────────────────
    for ri, day in enumerate(days):
        row_idx = hdr_row + 1 + ri
        ws.row_dimensions[row_idx].height = 52 if is_master else 38

        # Day label cell
        day_cell = ws.cell(row=row_idx, column=1, value=day)
        day_cell.font      = _font(bold=True, size=9, color=C_CODE_FG)
        day_cell.fill      = _fill(C_DAY_BG)
        day_cell.alignment = _align(wrap=False, halign="center", valign="center")
        day_cell.border    = _border()

        # Period content cells
        for ci, period in enumerate(periods, start=2):
            item       = grid.get(day, {}).get(period["id"])
            text       = _cell_text(item, is_master)
            cell       = ws.cell(row=row_idx, column=ci, value=text)
            cell.font  = _font(
                size=8,
                color=C_CODE_FG if (item and item != "–") else C_DET_FG,
            )
            cell.fill      = _fill(C_CELL_BG if ri % 2 == 0 else C_ALT_BG)
            cell.alignment = _align(wrap=True, halign="left", valign="top")
            cell.border    = _border()

    # ── Freeze pane below header + right of day col ────────────────────────
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=2)

    # ── Auto-filter on header row ─────────────────────────────────────────
    if periods:
        ws.auto_filter.ref = (
            f"A{hdr_row}:{ws.cell(row=hdr_row, column=1+len(periods)).column_letter}{hdr_row}"
        )

    # ── Print settings ────────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A3
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows       = f"$1:${hdr_row}"


# ─────────────────────────────────────────────────────────────────────────────
# Django response wrapper
# ─────────────────────────────────────────────────────────────────────────────

def _xlsx_response(data: bytes, filename: str) -> HttpResponse:
    """
    Wrap raw xlsx bytes in a Django HttpResponse with correct headers.

    Args:
        data     : bytes returned by _make_workbook()
        filename : e.g. "master_Sem1_2025.xlsx"
    """
    resp = HttpResponse(
        data,
        content_type=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp["Content-Length"]      = len(data)
    return resp
